import logging
from typing import Any, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side

from invoice_generator.core.config import sanitize_sheet_name

logger = logging.getLogger(__name__)

# ─── Layout Constants ───
COLUMN_WIDTHS = {
    "A": 11.57,
    "B": 10.29,
    "C": 7.29,
    "D": 10.29,
    "E": 11.29,
    "F": 24.14,
    "G": 15.43,
    "H": 10.57,
}
TOTAL_COLUMNS = 8
ROW_HEIGHT_COMPANY_NAME = 58.5
ROW_HEIGHT_DEFAULT = 13.5
ROW_HEIGHT_BUYER_INFO = 51.0
ROW_HEIGHT_BANK_DETAILS = 57.75
ROW_HEIGHT_TERMS_SECOND = 12.0
ROW_HEIGHT_TERMS_THIRD = 21.75
SPACER_BASE_HEIGHT = 304.5
SPACER_BASELINE_ROWS = 9
SPACER_MIN_HEIGHT = 10.0
MATERIAL_START_ROW = 10

# ─── Cell Formatting ───
CURRENCY_FORMAT = "[$₹]#,##0.00"
PERCENTAGE_FORMAT = "0.00%"
NUMBER_FORMAT = "#,##0.00"

BORDER_ALL = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
BORDER_SIDES = Border(left=Side(style="thin"), right=Side(style="thin"))
BORDER_SIDES_TOP = Border(
    left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin")
)
BORDER_SIDES_BOTTOM = Border(
    left=Side(style="thin"), right=Side(style="thin"), bottom=Side(style="thin")
)


def style_cell(
    cell,
    font_name: str = "Arial",
    font_size: int = 9,
    bold: bool = False,
    horizontal: str = "center",
    vertical: str = "center",
    wrap: bool = True,
    border: Optional[Border] = None,
    number_format: Optional[str] = None,
) -> None:
    """Apply font, alignment, border, and number format to an Excel cell."""
    cell.font = Font(name=font_name, size=font_size, bold=bold)
    cell.alignment = Alignment(horizontal=horizontal, vertical=vertical, wrap_text=wrap)
    if border:
        cell.border = border
    if number_format:
        cell.number_format = number_format


def apply_merged_row_borders(
    worksheet, row: int, total_cols: int = TOTAL_COLUMNS
) -> None:
    """Apply proper borders to a merged row spanning all columns."""
    worksheet.cell(row=row, column=1).border = Border(
        left=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin")
    )
    for column in range(2, total_cols):
        worksheet.cell(row=row, column=column).border = Border(
            top=Side(style="thin"), bottom=Side(style="thin")
        )
    worksheet.cell(row=row, column=total_cols).border = Border(
        right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin")
    )


def apply_side_borders(
    worksheet,
    row: int,
    start_col: int = 1,
    end_col: int = TOTAL_COLUMNS,
    border: Border = BORDER_SIDES,
) -> None:
    for column in range(start_col, end_col + 1):
        worksheet.cell(row=row, column=column).border = border


def write_tax_line(
    worksheet,
    row: int,
    label: str,
    rate: Optional[float],
    amount: float,
    border: Border = BORDER_SIDES,
) -> None:
    apply_side_borders(worksheet, row, end_col=TOTAL_COLUMNS - 1)
    cell = worksheet.cell(row=row, column=5, value=label)
    style_cell(cell, bold=(label == "TOTAL"), border=border)
    if rate is not None:
        cell = worksheet.cell(row=row, column=6, value=rate)
        style_cell(cell, font_size=10, number_format=PERCENTAGE_FORMAT, border=border)
    cell = worksheet.cell(row=row, column=8, value=amount)
    style_cell(
        cell,
        font_size=10,
        bold=(label == "TOTAL"),
        number_format=CURRENCY_FORMAT,
        border=border,
    )


class InvoiceExcelRenderer:
    """Renderer responsible for translating processed data dictionaries into formatted Excel Tax Workbooks."""

    def __init__(self, data: dict[str, Any]):
        self.data = data
        self.config = data["config"]

    def render(self, workbook: Workbook) -> None:
        """Fully construct and configure the workbook with all necessary sheets."""
        workbook.remove(workbook.active)
        self._create_invoice_sheet(workbook)
        self._create_list_sheet(workbook)

    def _create_invoice_sheet(self, workbook: Workbook) -> None:
        sheet_name = sanitize_sheet_name(f"Bill - {self.data['inv_num']}")
        worksheet = workbook.create_sheet(title=sheet_name)

        num_material_rows = len(self.data["sites_data"]) * len(self.data["materials"])

        for letter, width in COLUMN_WIDTHS.items():
            worksheet.column_dimensions[letter].width = width

        # Row Assignments
        material_end_row = MATERIAL_START_ROW + num_material_rows - 1
        spacer_row = material_end_row + 1
        total_row = spacer_row + 1
        cgst_row = total_row + 1
        sgst_row = cgst_row + 1
        roundoff_row = sgst_row + 1
        grand_total_row = roundoff_row + 1
        bank_row = grand_total_row + 1
        terms_header_row = bank_row + 1
        terms_row_2 = terms_header_row + 1
        terms_row_3 = terms_row_2 + 1

        worksheet.row_dimensions[1].height = ROW_HEIGHT_COMPANY_NAME
        for row in range(2, 8):
            worksheet.row_dimensions[row].height = ROW_HEIGHT_DEFAULT
        worksheet.row_dimensions[8].height = ROW_HEIGHT_BUYER_INFO
        worksheet.row_dimensions[9].height = ROW_HEIGHT_DEFAULT

        spacer_height = max(
            SPACER_MIN_HEIGHT,
            SPACER_BASE_HEIGHT
            - (num_material_rows - SPACER_BASELINE_ROWS) * ROW_HEIGHT_DEFAULT,
        )
        worksheet.row_dimensions[spacer_row].height = spacer_height

        for row in [total_row, cgst_row, sgst_row, roundoff_row, grand_total_row]:
            worksheet.row_dimensions[row].height = ROW_HEIGHT_DEFAULT
        worksheet.row_dimensions[bank_row].height = ROW_HEIGHT_BANK_DETAILS
        worksheet.row_dimensions[terms_header_row].height = ROW_HEIGHT_DEFAULT
        worksheet.row_dimensions[terms_row_2].height = ROW_HEIGHT_TERMS_SECOND
        worksheet.row_dimensions[terms_row_3].height = ROW_HEIGHT_TERMS_THIRD

        self._write_company_header(worksheet)
        self._write_buyer_info(worksheet)
        self._write_column_headers(worksheet)
        self._write_materials(worksheet, current_row=MATERIAL_START_ROW)

        apply_side_borders(worksheet, spacer_row)

        self._write_totals(
            worksheet, total_row, cgst_row, sgst_row, roundoff_row, grand_total_row
        )
        self._write_bank_and_terms(
            worksheet, bank_row, terms_header_row, terms_row_2, terms_row_3
        )

    def _write_company_header(self, worksheet):
        company = self.config["company"]
        header_lines = [
            (company["name"], 41, False),
            (company["subtitle"], 9, False),
            (company["address"], 9, False),
            (f"Contact No: {company['contact']}", 9, False),
            (f"GSTN {company['gstn']}", 9, False),
            (f"PAN NO {company['pan']}", 9, False),
            ("TAX INVOICE", 9, True),
        ]
        for row_index, (text, font_size, is_bold) in enumerate(header_lines, start=1):
            worksheet.merge_cells(f"A{row_index}:H{row_index}")
            cell = worksheet.cell(row=row_index, column=1, value=text)
            style_cell(cell, font_size=font_size, bold=is_bold)
            apply_merged_row_borders(worksheet, row_index)

    def _write_buyer_info(self, worksheet):
        buyer = self.config["buyer"]
        buyer_text = f"{buyer['name']}\n{buyer['address'].strip()}\nBUYER - GSTIN {buyer['gstn']}"
        worksheet.merge_cells("A8:D8")
        cell = worksheet.cell(row=8, column=1, value=buyer_text)
        style_cell(cell, bold=True)
        worksheet.cell(row=8, column=1).border = BORDER_ALL
        for col in range(2, 4):
            worksheet.cell(row=8, column=col).border = Border(
                top=Side(style="thin"), bottom=Side(style="thin")
            )
        worksheet.cell(row=8, column=4).border = Border(
            right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin")
        )

        worksheet.merge_cells("E8:F8")
        cell = worksheet.cell(
            row=8,
            column=5,
            value=f"INVOICE NO\n{self.data['inv_num']}-{self.data['fiscal_year']}",
        )
        style_cell(cell, bold=True)
        worksheet.cell(row=8, column=5).border = BORDER_ALL
        worksheet.cell(row=8, column=6).border = Border(
            right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin")
        )

        worksheet.merge_cells("G8:H8")
        cell = worksheet.cell(
            row=8,
            column=7,
            value=f"DATE: {self.data['invoice_date'].strftime('%d-%m-%Y')}",
        )
        style_cell(cell, bold=True)
        worksheet.cell(row=8, column=7).border = BORDER_ALL
        worksheet.cell(row=8, column=8).border = Border(
            right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin")
        )

    def _write_column_headers(self, worksheet):
        headers = [
            "Date",
            "Material",
            "Site",
            "Challan",
            "Quantity",
            "Rate",
            "Per",
            "Amount",
        ]
        for col_index, header_text in enumerate(headers, start=1):
            cell = worksheet.cell(row=9, column=col_index, value=header_text)
            style_cell(cell, bold=True, border=BORDER_ALL)

    def _write_materials(self, worksheet, current_row: int):
        for site_info in self.data["sites_data"]:
            first_row = current_row
            last_row = current_row + len(self.data["materials"]) - 1

            if last_row > first_row:
                worksheet.merge_cells(f"A{first_row}:A{last_row}")
                worksheet.merge_cells(f"D{first_row}:D{last_row}")

            cell = worksheet.cell(
                row=first_row, column=1, value=site_info["date_range_label"]
            )
            style_cell(cell, font_size=10, border=BORDER_SIDES_TOP)
            cell = worksheet.cell(row=first_row, column=4, value="As Per List")
            style_cell(cell, border=BORDER_SIDES_TOP)

            for material_index, material_name in enumerate(self.data["materials"]):
                row = current_row + material_index
                material_info = site_info["material_data"].get(material_name, {})
                qty = material_info.get("qty")
                rate = material_info.get("rate")
                is_first = material_index == 0

                b_style = BORDER_SIDES_TOP if is_first else BORDER_SIDES

                worksheet.cell(row=row, column=2, value=material_name)
                style_cell(worksheet.cell(row=row, column=2), border=b_style)

                worksheet.cell(row=row, column=3, value=site_info["site_name"])
                style_cell(worksheet.cell(row=row, column=3), border=b_style)

                if qty is not None:
                    worksheet.cell(row=row, column=5, value=round(qty, 2))
                style_cell(worksheet.cell(row=row, column=5), border=b_style)

                if rate is not None:
                    worksheet.cell(row=row, column=6, value=float(rate))
                style_cell(worksheet.cell(row=row, column=6), border=b_style)

                worksheet.cell(row=row, column=7, value=self.data["unit"])
                style_cell(worksheet.cell(row=row, column=7), border=b_style)

                amount = (
                    round(qty, 2) * float(rate)
                    if qty is not None and rate is not None
                    else None
                )
                worksheet.cell(row=row, column=8, value=amount)
                style_cell(
                    worksheet.cell(row=row, column=8), font_size=10, border=b_style
                )

                if not is_first:
                    worksheet.cell(row=row, column=1).border = BORDER_SIDES
                    worksheet.cell(row=row, column=4).border = BORDER_SIDES

            current_row = last_row + 1

    def _write_totals(
        self, worksheet, total_row, cgst_row, sgst_row, roundoff_row, grand_total_row
    ):
        write_tax_line(worksheet, total_row, "TOTAL", None, self.data["total_amount"])
        write_tax_line(
            worksheet,
            cgst_row,
            "CGST",
            self.config["gst"]["cgst_rate"],
            self.data["cgst_amount"],
        )
        write_tax_line(
            worksheet,
            sgst_row,
            "SGST",
            self.config["gst"]["sgst_rate"],
            self.data["sgst_amount"],
        )

        cell = worksheet.cell(row=sgst_row, column=2, value="HSN Code")
        style_cell(cell, font_size=10, border=BORDER_SIDES)
        cell = worksheet.cell(
            row=sgst_row, column=3, value=float(self.config["gst"]["hsn_code"])
        )
        style_cell(cell, font_size=10, border=BORDER_SIDES)

        apply_side_borders(
            worksheet,
            roundoff_row,
            end_col=TOTAL_COLUMNS - 1,
            border=BORDER_SIDES_BOTTOM,
        )
        cell = worksheet.cell(row=roundoff_row, column=5, value="ROUND OFF")
        style_cell(cell, border=BORDER_SIDES_BOTTOM)
        cell = worksheet.cell(row=roundoff_row, column=8, value=self.data["round_off"])
        style_cell(
            cell,
            font_size=10,
            number_format=CURRENCY_FORMAT,
            border=BORDER_SIDES_BOTTOM,
        )

        worksheet.merge_cells(f"A{grand_total_row}:G{grand_total_row}")
        cell = worksheet.cell(row=grand_total_row, column=1, value="GRAND TOTAL")
        style_cell(cell, bold=True, border=BORDER_ALL)
        for column in range(2, 8):
            worksheet.cell(row=grand_total_row, column=column).border = Border(
                top=Side(style="thin"),
                bottom=Side(style="thin"),
                right=Side(style="thin") if column == 7 else None,
            )
        cell = worksheet.cell(
            row=grand_total_row, column=8, value=self.data["grand_total"]
        )
        style_cell(cell, font_size=10, number_format=CURRENCY_FORMAT, border=BORDER_ALL)

    def _write_bank_and_terms(
        self, worksheet, bank_row, terms_header_row, terms_row_2, terms_row_3
    ):
        bank = self.config["bank"]
        bank_text = f"BANK DETAILS :- {bank['account_name']} BANK NAME: {bank['bank_name']}\nA/C NO: {bank['account_no']}\nBRANCH & IFSC : {bank['branch']} / {bank['ifsc']}"
        worksheet.merge_cells(f"A{bank_row}:E{bank_row}")
        cell = worksheet.cell(row=bank_row, column=1, value=bank_text)
        style_cell(cell, bold=True)
        worksheet.cell(row=bank_row, column=1).border = BORDER_ALL
        for col in range(2, 5):
            worksheet.cell(row=bank_row, column=col).border = Border(
                top=Side(style="thin"), bottom=Side(style="thin")
            )
        worksheet.cell(row=bank_row, column=5).border = Border(
            right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin")
        )

        worksheet.merge_cells(f"F{bank_row}:H{bank_row}")
        cell = worksheet.cell(
            row=bank_row, column=6, value=self.config["company"]["name"]
        )
        style_cell(cell, font_size=10, bold=True)
        worksheet.cell(row=bank_row, column=6).border = BORDER_ALL
        for col in range(7, 9):
            worksheet.cell(row=bank_row, column=col).border = Border(
                top=Side(style="thin"),
                bottom=Side(style="thin"),
                right=Side(style="thin") if col == 8 else None,
            )

        worksheet.merge_cells(f"A{terms_header_row}:B{terms_header_row}")
        worksheet.merge_cells(f"C{terms_header_row}:H{terms_header_row}")
        cell = worksheet.cell(
            row=terms_header_row, column=1, value="Terms & Conditions :"
        )
        style_cell(cell)
        for col in range(1, 9):
            worksheet.cell(row=terms_header_row, column=col).border = BORDER_ALL

        terms_content = [
            (
                terms_row_2,
                "1) interest @ 18% will be charged on payment due for more than 30 days from date of invoice.",
                7,
            ),
            (
                terms_row_3,
                "2) Our responsibility ceases if any error is not reported in writing 7 days from date of invoice.",
                7,
            ),
            (
                terms_row_3 + 1,
                '"Input tax Credit of CGST / SGST / IGST charged on goods and services used exclusively or partly in supplying goods transport agency services has not been taken."',
                10,
            ),
        ]
        for row_number, text, font_size in terms_content:
            worksheet.merge_cells(f"A{row_number}:H{row_number}")
            cell = worksheet.cell(row=row_number, column=1, value=text)
            style_cell(cell, font_size=font_size)
            for col in range(1, 9):
                worksheet.cell(row=row_number, column=col).border = BORDER_ALL

    def _create_list_sheet(self, workbook: Workbook) -> None:
        sheet_name = sanitize_sheet_name(f"List - {self.data['inv_num']}")
        worksheet = workbook.create_sheet(title=sheet_name)
        headers = [
            "Date",
            "Challan No.",
            "Vehicle No.",
            "Site",
            "Material",
            "Quantity",
            "Rate",
            "Per",
            "Amount",
        ]
        for col_index, header_text in enumerate(headers, start=1):
            cell = worksheet.cell(row=1, column=col_index, value=header_text)
            style_cell(cell, bold=True, border=BORDER_ALL)

        df = self.data["df"]
        if df.empty:
            return

        for col, width in enumerate([12, 12, 15, 20, 15, 10, 10, 10, 12], start=1):
            worksheet.column_dimensions[
                worksheet.cell(row=1, column=col).column_letter
            ].width = width

        # Calculate totals
        quantity_total = 0.0
        amount_total = 0.0

        for row_offset, (_, row_data) in enumerate(df.iterrows()):
            row_number = row_offset + 2

            import pandas as pd

            def safe_float(v):
                return float(v) if pd.notna(v) and str(v).strip() != "" else None

            qty = safe_float(row_data.get("Quantity")) or 0.0
            rate = safe_float(row_data.get("Rate")) or 0.0
            amount = qty * rate

            quantity_total += qty
            amount_total += amount

            per_unit = row_data.get("Per", "") if pd.notna(row_data.get("Per")) else ""

            values = [
                str(row_data.get("Date", ""))[:10]
                if pd.notna(row_data.get("Date"))
                else "",
                row_data.get("Challan No.", ""),
                row_data.get("Vehicle No.", ""),
                row_data.get("Site", ""),
                row_data.get("Material", ""),
                qty,
                rate,
                per_unit,
                amount,
            ]

            for column_index, value in enumerate(values, start=1):
                cell = worksheet.cell(row=row_number, column=column_index, value=value)
                style_cell(cell, border=BORDER_SIDES)
                if column_index in (6, 7, 9):
                    cell.number_format = NUMBER_FORMAT

        totals_row = len(df) + 2
        cell = worksheet.cell(row=totals_row, column=6, value=quantity_total)
        style_cell(cell, border=BORDER_ALL, number_format=NUMBER_FORMAT)

        cell = worksheet.cell(row=totals_row, column=9, value=amount_total)
        style_cell(cell, border=BORDER_ALL, number_format=NUMBER_FORMAT)
